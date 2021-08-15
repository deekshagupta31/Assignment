# Imports 
import scrapy
import pandas as pd
from scrapy.crawler import CrawlerProcess
from openpyxl import Workbook
from openpyxl.styles import Font
from yattag import Doc, indent

# Instantiate a new workbook
wb = Workbook()
# Choose the active worksheet in the workbook
ws = wb.active
# Added column Name in the sheet and make them Bold
columnName=["Book Name","Book Price","Book Avaliablity","Book Image Path","Book Star Rating(From 5)"]
ws.append(columnName)
ws['A1'].font = Font(bold=True)
ws['B1'].font = Font(bold=True)
ws['C1'].font = Font(bold=True)
ws['D1'].font = Font(bold=True)
ws['E1'].font = Font(bold=True)
class WebSpider(scrapy.Spider):
    name = "WebSpider"    
    start_urls = ["http://books.toscrape.com/"]
    def parse(self,webPage):        
        #theCards will a List containing HTML elements
        theCards = webPage.css('html.no-js body#default.default div.container-fluid.page div.page_inner div.row div.col-sm-8.col-md-9 section div ol.row li.col-xs-6.col-sm-4.col-md-3.col-lg-3')        
        # Loop through all of the cards        
        for card in theCards:
            # Create an empty list for adding data for each book
            excelRow = []
            # name of the book in the excelRow
            name = card.css('article.product_pod h3 a ::attr(title)').get()           
            excelRow.append(name)
            # price of the book
            price = card.css('article.product_pod div.product_price p.price_color ::text').get()
            excelRow.append(price)
            # avaliability of the book
            avaliableList =card.css('article.product_pod div.product_price p.instock.availability ::text').getall()
            avaliability=avaliableList[1].replace("\\n","").strip()
            excelRow.append(avaliability)
            # image path of the book
            image = card.css('article.product_pod div.image_container a ::attr(src)').get()
            excelRow.append(image)
            # stars for the book
            stars=card.css('article.product_pod p ::attr(class)').get().replace("star-rating ", "")                    
            excelRow.append(stars)
            # Add the excelRow to the worksheet
            ws.append(excelRow)
        # end card looped for each book    
        # Example for paginating by clicking on the next page
        for next_page in webPage.css('html.no-js body#default.default div.container-fluid.page div.page_inner div.row div.col-sm-8.col-md-9 section div div ul.pager li.next a'):
            yield webPage.follow(next_page, self.parse)
        # Save the whole workbook       
        wb.save('books.xlsx')
        # Save Workbook as CVS also
        read_file = pd.read_excel (r'books.xlsx')
        read_file.to_csv (r'books_csv.csv', index = None, header=True)
        #convert workbook to xml
        self.convertExcelToXml()

    def convertExcelToXml(self):        
        # Returning returns a triplet
        doc, tag, text = Doc().tagtext()
        with tag('BooksForSale'):
            for row in ws.iter_rows(min_row=2, max_row=10, min_col=1, max_col=6):
                row = [cell.value for cell in row]
                with tag("Books"):
                    with tag("Book_Name"):
                        text(row[0])
                    with tag("Book_Price"):
                        text(row[1])
                    with tag("Book_Avaliablity"):
                        text(row[2])
                    with tag("Book_Image_Path"):
                        text(row[3])
                    with tag("Book_Rating"):
                        text(row[4])                    
                #end book tag
            #end for loop
        #end BooksForSale tag
        result = indent(doc.getvalue(),indentation='   ',indent_text=True)
        with open("booksForSal.xml", "w") as f:
            f.write(result)
    #end convertExcelToXml method        
# Instantiate and run our WebSpider
process = CrawlerProcess()
process.crawl(WebSpider)
process.start()
